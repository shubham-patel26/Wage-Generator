from flask import Flask, render_template, request, send_file
import pandas as pd
import pdfplumber
import re
import os
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), '..', 'uploads')
app.config['OUTPUT_FOLDER'] = os.path.join(os.path.dirname(__file__), '..', 'output')

def extract_formb_with_uan(file_path):
    df = pd.read_excel(file_path)
    sr_col = next((col for col in df.columns if 'sr' in col.lower()), None)
    uan_col = next((col for col in df.columns if 'uan' in col.lower()), None)
    days_col = next((col for col in df.columns if 'day' in col.lower()), None)
    
    df[uan_col] = pd.to_numeric(df[uan_col], errors='coerce').fillna(0).astype(int).astype(str).str.zfill(12).str.strip()
    filtered = df[df[days_col] > 0]
    
    return filtered[[sr_col, uan_col]].values.tolist()

def extract_pf_pdf_to_excel(pdf_path, output_dir):
    all_rows = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages[3:]:  # Skip first 3 pages
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if not row:
                        continue
                    
                    # Look for UAN in the row
                    uan_match = None
                    for cell in row:
                        if cell and re.search(r'\b\d{12}\b', str(cell).strip()):
                            uan_match = re.search(r'\b\d{12}\b', str(cell).strip()).group(0)
                            break
                    
                    if uan_match:
                        # Clean and pad row to consistent length
                        clean_row = [str(cell).strip() if cell else '' for cell in row]
                        all_rows.append(clean_row)
    
    if all_rows:
        # Create DataFrame with dynamic columns based on max row length
        max_cols = max(len(row) for row in all_rows) if all_rows else 0
        columns = [f'Col_{i+1}' for i in range(max_cols)]
        
        # Pad all rows to same length
        padded_rows = [row + [''] * (max_cols - len(row)) for row in all_rows]
        
        df = pd.DataFrame(padded_rows, columns=columns)
        
        # Add UAN column for easy identification
        df['UAN'] = df.apply(lambda row: next((re.search(r'\b\d{12}\b', str(cell)).group(0) 
                                              for cell in row if cell and re.search(r'\b\d{12}\b', str(cell))), ''), axis=1)
        
        excel_path = os.path.join(output_dir, 'PF_Member_Details.xlsx')
        df.to_excel(excel_path, index=False, engine='openpyxl')
        
        return df, excel_path
    else:
        # Return empty DataFrame if no data found
        df = pd.DataFrame()
        excel_path = os.path.join(output_dir, 'PF_Member_Details.xlsx')
        df.to_excel(excel_path, index=False)
        return df, excel_path

def match_and_highlight_excel(formb_data, pf_df, excel_path):
    # Create UAN to row mapping
    formb_uans = {uan: sr_no for sr_no, uan in formb_data}
    
    matched_rows = []
    matches = 0
    unmatched = 0
    
    # Load workbook and get active sheet
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Define highlight color
    highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    fillColorTillColumn = ws.max_column + 1
    addSROnColumn = fillColorTillColumn + 1
    # Check each row in PF data
    for idx, row in pf_df.iterrows():
        uan = row['UAN']
        if uan in formb_uans:
            # Highlight the row in Excel (idx + 2 because of header and 0-based index)
            excel_row = idx + 2
            for col in range(1, fillColorTillColumn):
                ws.cell(row=excel_row, column=col).fill = highlight_fill
            
            # Add serial number in a new column
            ws.cell(row=excel_row, column=addSROnColumn, value=f"SR-{formb_uans[uan]}")
            
            matched_rows.append(excel_row)
            matches += 1
        else:
            unmatched += 1
    
    # Add header for serial number column
    ws.cell(row=1, column=ws.max_column, value="Form_B_SR_No")
    
    # Save the highlighted Excel
    highlighted_path = excel_path.replace('.xlsx', '_Highlighted.xlsx')
    wb.save(highlighted_path)
    
    return matches, unmatched, matched_rows, highlighted_path

@app.route('/', methods=['GET', 'POST'])
def index():
    message = None
    download_link = None
    
    if request.method == 'POST':
        formb_files = request.files.getlist('formb_files')
        pf_file = request.files.get('pf_file')
        font_size = int(request.form.get('font_size', 10))
        
        if not formb_files or not pf_file:
            message = 'Please select both Form B Excel files and PF PDF file.'
            return render_template('index.html', message=message)
        
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)
        
        pf_filename = secure_filename(pf_file.filename)
        pf_path = os.path.join(app.config['UPLOAD_FOLDER'], pf_filename)
        pf_file.save(pf_path)
        
        pf_df, excel_path = extract_pf_pdf_to_excel(pf_path, app.config['OUTPUT_FOLDER'])
        summary_messages = []
        
        for formb_file in formb_files:
            if formb_file.filename:
                formb_filename = secure_filename(formb_file.filename)
                formb_path = os.path.join(app.config['UPLOAD_FOLDER'], formb_filename)
                formb_file.save(formb_path)
                
                formb_filtered = extract_formb_with_uan(formb_path)
                total_workers = len(formb_filtered)
                
                if not formb_filtered:
                    summary_messages.append(f"{formb_filename} → No valid workers.")
                    continue
                
                matches, unmatched, matched_rows, highlighted_excel = match_and_highlight_excel(formb_filtered, pf_df, excel_path)
                
                formb_name = os.path.splitext(formb_filename)[0]
                summary_messages.append(f"{formb_name} → Total: {total_workers}, Matched: {matches}, Unmatched: {unmatched}")
                summary_messages.append(f"Matched rows: {', '.join(map(str, matched_rows))}")
                download_link = os.path.basename(highlighted_excel)
        
        message = '\n'.join(summary_messages)
    
    return render_template('index.html', message=message, download_link=download_link)

@app.route('/download/<filename>')
def download_file(filename):
    return send_file(os.path.join(app.config['OUTPUT_FOLDER'], filename), as_attachment=True)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080, debug=True)